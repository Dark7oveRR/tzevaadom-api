import requests
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict

def saveFn():
    url = "https://api.tzevaadom.co.il/alerts-history/"
    response = requests.get(url)
    
    if response.status_code != 200:
        print(f"Failed to fetch data. Status code: {response.status_code}")
        return
    
    data = response.json()
    
    city_count = defaultdict(int)
    city_threat = defaultdict(int)

    for item in data:
        if "alerts" in item and isinstance(item["alerts"], list):
            for alert in item["alerts"]:
                cities = set(alert.get("cities", []))
                for city in cities:
                    city_count[city] += 1
                    city_threat[city] += alert.get("threat", 0)
    
    sorted_city_analysis = sorted(city_threat.items(), key=lambda x: x[1], reverse=True)
    
    wb = openpyxl.Workbook()
    ws_cities = wb.active
    ws_cities.title = "City Analysis"
    ws_cities.append(["City", "Target Count", "Total Threat Level"])
    
    for city, threat in sorted_city_analysis:
        ws_cities.append([city, city_count[city], threat])
    
    for col in ws_cities.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = get_column_letter(col[0].column)
        ws_cities.column_dimensions[col_letter].width = max_length + 5
    
    excel_file = "City_Analysis.xlsx"
    wb.save(excel_file)
    print(f"City analysis data saved to {excel_file}")

saveFn()
